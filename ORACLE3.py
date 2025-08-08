import asyncio
from asyncua import Client, ua
from datetime import datetime, timedelta, time
import cx_Oracle
from openpyxl import load_workbook

# -------------------- Configuración OPC UA --------------------
URL = "opc.tcp://localhost:52250/freeopcua/server/"
NODE_GROUPS = ["BOMBAS_SH", "CONVEYOR_SH"] 
START_HEX = 0x80
END_HEX = 0xFF
INTERVALO_SEGUNDOS = 0.05

# -------------------- Archivos Excel para descripción --------------------
ARCHIVOS_DESCRIPCIONES = {
    "BOMBAS_SH": r"GMFs_de_Equipos\GMF_BOMBAS_SH.xlsx",
    "CONVEYOR_SH": r"GMFs_de_Equipos\GMF_CONVEYOR_SH.xlsx",
}

# -------------------- Variables de estado --------------------
tiempos_gmf = {}
tiempos_congelados = {}
tiempos_continuous = {}
acumulado_continuous = {}
ultima_marca_continuous = {}
descripciones_gmf = {}
pending_export = {}
turno_registrado = set()

# -------------------- Determinar turno de producción --------------------
def determinar_turno(fecha: datetime) -> str:
    hora = fecha.time()
    if time(6, 5) <= hora <= time(14, 30):
        return "MAÑANA"
    elif time(14, 39) <= hora <= time(22, 54):
        return "TARDE"
    elif time(23, 13) <= hora <= time(23, 59) or time(0, 0) <= hora <= time(5, 55):
        return "NOCHE"
    return "FUERA DE TURNO"

# -------------------- Validar inicio de turno en Oracle --------------------
def ya_registrado_en_oracle(grupo, turno, inicio):
    try:
        connection = cx_Oracle.connect(user='HR', password='12345', dsn='localhost:1521/orcl')
        cursor = connection.cursor()
        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        cursor.execute(f"""
            SELECT COUNT(*) FROM {tabla}
            WHERE TURNO = :1 AND FALLA = '--' AND TRUNC(INICIO_FALLA) = TRUNC(:2)
        """, (turno, inicio))
        count, = cursor.fetchone()
        return count > 0
    except Exception as e:
        print(f"❌ Error validando inicio de turno en Oracle: {e}")
        return False
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

# -------------------- Mostrar inicio de turno actual --------------------
def mostrar_inicio_turno():
    ahora = datetime.now()
    turno = determinar_turno(ahora)
    if turno == "MAÑANA":
        inicio = ahora.replace(hour=6, minute=5, second=0, microsecond=0)
    elif turno == "TARDE":
        inicio = ahora.replace(hour=14, minute=39, second=0, microsecond=0)
    elif turno == "NOCHE":
        if ahora.hour >= 23:
            inicio = ahora.replace(hour=23, minute=13, second=0, microsecond=0)
        else:
            ayer = ahora - timedelta(days=1)
            inicio = ayer.replace(hour=23, minute=13, second=0, microsecond=0)
    else:
        inicio = ahora
    print(f"📌 Inicio del turno {turno}: {inicio.strftime('%Y-%m-%d %H:%M:%S')}")



# -------------------- Cargar descripciones desde Excel --------------------
def cargar_descripciones():
    for grupo, ruta in ARCHIVOS_DESCRIPCIONES.items():
        descripciones_gmf[grupo] = {}
        try:
            wb = load_workbook(ruta)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo, descripcion = fila[:2]
                if codigo and descripcion:
                    descripciones_gmf[grupo][codigo.strip().upper()] = descripcion.strip()
        except Exception as e:
            print(f"❌ Error al cargar descripciones de {grupo}: {e}")

# -------------------- Exportar a Oracle --------------------
def exportar_a_oracle(grupo, gmf, descripcion, inicio, fin, acumulado):
    try:
        connection = cx_Oracle.connect(
            user='HR',
            password='12345',
            dsn='localhost:1521/orcl',
        )
        cursor = connection.cursor()

        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        turno = determinar_turno(inicio)

        cursor.execute(f"""
            BEGIN
                EXECUTE IMMEDIATE '
                    CREATE TABLE {tabla} (
                        EQUIPO VARCHAR2(50),
                        FALLA VARCHAR2(50),
                        DESCRIPCION VARCHAR2(255),
                        INICIO_FALLA TIMESTAMP,
                        FIN_FALLA TIMESTAMP,
                        TIEMPO_OFF_SEG NUMBER,
                        TURNO VARCHAR2(20)
                    )';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -955 THEN RAISE; END IF;
            END;
        """)

        cursor.execute(f"""
            INSERT INTO {tabla} (EQUIPO, FALLA, DESCRIPCION, INICIO_FALLA, FIN_FALLA, TIEMPO_OFF_SEG, TURNO)
            VALUES (:1, :2, :3, :4, :5, :6, :7)
        """, (grupo, gmf, descripcion, inicio, fin, acumulado, turno))

        connection.commit()
        print(f"📃 Exportado a Oracle: {grupo} | {gmf} | {turno} | {acumulado} seg")

    except Exception as e:
        print(f"❌ Error al exportar a Oracle: {e}")
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

# -------------------- Lectura por grupo --------------------
async def leer_grupo_si_falla(client, grupo):
    ahora = datetime.now()

    try:
        # Leer TOTAL_START del grupo CONVEYOR_SH
        nodo_total_start_global = client.get_node("ns=2;s=CONVEYOR_SH.TOTAL_START")
        total_start_value = await nodo_total_start_global.read_value()
        if not total_start_value:
            print(f"⏹ TOTAL_START está en OFF, omitiendo monitoreo para todos los grupos...")
            await asyncio.sleep(0.5)
            return

        nodo_fault = client.get_node(f"ns=2;s={grupo}.TOTAL_FAULT")
        nodo_cont = client.get_node(f"ns=2;s={grupo}.CONTINUOUS")
        valor_fault, valor_continuous = await client.read_values([nodo_fault, nodo_cont])

        if grupo not in acumulado_continuous:
            acumulado_continuous[grupo] = timedelta(0)
            ultima_marca_continuous[grupo] = ahora

        if not valor_continuous:
            delta = ahora - ultima_marca_continuous[grupo]
            acumulado_continuous[grupo] += delta
        ultima_marca_continuous[grupo] = ahora

        NODE_PREFIX = f"ns=2;s={grupo}.GMF"
        hex_suffixes = [f"{i:X}" for i in range(START_HEX, END_HEX + 1)]
        node_ids = [ua.NodeId.from_string(f"{NODE_PREFIX}{suf}") for suf in hex_suffixes]
        nodes = [client.get_node(nid) for nid in node_ids]
        values = await client.read_values(nodes)

        gmfs_detectadas = set()

        for suffix, value in zip(hex_suffixes, values):
            if isinstance(value, int):
                for bit_pos in range(16):
                    if (value >> bit_pos) & 1:
                        gmf_base = f"{suffix}{bit_pos:X}"
                        gmf_id = f"{grupo}.GMF{gmf_base.upper()}"
                        gmf_solo = f"GMF{gmf_base.upper()}"
                        gmfs_detectadas.add(gmf_id)

                        if gmf_id not in tiempos_gmf:
                            tiempos_gmf[gmf_id] = ahora

                        segundos = int((ahora - tiempos_gmf[gmf_id]).total_seconds())
                        tiempos_congelados[gmf_id] = segundos

                        descripcion = descripciones_gmf.get(grupo, {}).get(gmf_solo.strip().upper(), "Sin descripción")
                        inicio_falla = tiempos_gmf[gmf_id].strftime("%Y-%m-%d %H:%M:%S")

                        acumulado = int(acumulado_continuous[grupo].total_seconds())
                        estado_cont = "OFF" if not valor_continuous else "ON"

                        print(f"⏱ {gmf_id} activa hace {segundos} seg → {descripcion} | Inicio de falla: {inicio_falla} | CONTINUOUS {estado_cont} → Acumulado: {acumulado} seg")
            else:
                print(f"{grupo}.GMF{suffix} → Valor no entero: {value}")

        inactivas = [gmf for gmf in tiempos_gmf if gmf.startswith(grupo) and gmf not in gmfs_detectadas]
        for gmf in inactivas:
            grupo_actual, gmf_codigo = gmf.split(".")
            nodo_gmf = client.get_node(f"ns=2;s={grupo}.{gmf_codigo}")
            val_fault, val_cont, val_gmf = await client.read_values([nodo_fault, nodo_cont, nodo_gmf])
            descripcion = descripciones_gmf.get(grupo, {}).get(gmf_codigo.strip().upper(), "Sin descripción")
            inicio_falla = tiempos_gmf.get(gmf)
            segundos = tiempos_congelados.get(gmf, 0)
            acumulado = int(acumulado_continuous[grupo].total_seconds())

            if not val_gmf and not val_cont:
                pending_export[gmf] = (grupo, gmf_codigo, descripcion, inicio_falla, ahora, acumulado)
            elif not val_gmf and val_cont:
                if gmf in pending_export:
                    grupo, gmf_codigo, descripcion, inicio, _, _ = pending_export.pop(gmf)
                    exportar_a_oracle(grupo, gmf_codigo, descripcion, inicio, ahora, acumulado)
                else:
                    exportar_a_oracle(grupo, gmf_codigo, descripcion, inicio_falla, ahora, acumulado)

                tiempos_gmf.pop(gmf, None)
                tiempos_congelados.pop(gmf, None)

                print(f"✅ Exportado: {grupo} | {gmf_codigo} | {descripcion} | {inicio_falla} → {ahora} | {acumulado} seg")

            estado_cont = "OFF" if not val_cont else "ON"
            linea = f"⏱ {gmf} inactiva desde hace {segundos} seg → {descripcion} | Inicio de falla: {inicio_falla} | CONTINUOUS {estado_cont} → Acumulado: {acumulado} seg"
            print(linea)

        if not any(k.startswith(grupo) for k in tiempos_gmf) and not any(k.startswith(grupo) for k in pending_export):
            acumulado_continuous[grupo] = timedelta(0)

    except Exception as e:
        print(f"❌ Error en grupo {grupo}: {e}")

# -------------------- Bucle principal --------------------
async def main():
    cargar_descripciones()
    mostrar_inicio_turno()
    async with Client(url=URL) as client:
        print("✅ Conectado al servidor OPC UA")
        while True:
            await asyncio.gather(*[leer_grupo_si_falla(client, grupo) for grupo in NODE_GROUPS])
            await asyncio.sleep(INTERVALO_SEGUNDOS)

# -------------------- Ejecutar --------------------
if __name__ == "__main__":
    asyncio.run(main())

